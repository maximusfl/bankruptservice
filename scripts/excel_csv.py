#!/usr/bin/env python3

import pandas as pd
import os
import logging
from openpyxl import load_workbook

logging.basicConfig(filename='excel_to_csv.log', level=logging.INFO,
                    format='%(asctime)s:%(levelname)s:%(message)s')

def preprocess_dataframe(df):
    return df


def parse_sheet(df, writer, sheet_name):

    df = preprocess_dataframe(df)

    writer.writerow([f'Sheet: {sheet_name}'])

    df.to_csv(writer, index=False, header=True)

    writer.writerow([])

def excel_to_csv(excel_file_path, csv_file_path):
    try:
        workbook = load_workbook(excel_file_path)
        sheet_names = workbook.sheetnames

        if os.path.exists(csv_file_path):
            logging.warning(f'The file {csv_file_path} already exists. It will be overwritten.')

        with open(csv_file_path, 'w', newline='') as csvfile:
            writer = csv.writer(csvfile)

            for sheet_name in sheet_names:
                df = pd.read_excel(excel_file_path, sheet_name=sheet_name)

                parse_sheet(df, writer, sheet_name)

        logging.info(f'Successfully converted {excel_file_path} to {csv_file_path}')
    except Exception as e:
        logging.error(f'Error occurred: {e}')

excel_file_path = input('Enter the path to the Excel file: ')
csv_file_path = input('Enter the desired path for the CSV file: ')

excel_to_csv(excel_file_path, csv_file_path)